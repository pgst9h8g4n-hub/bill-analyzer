"""
微信 Excel 账单解析器
前17行元数据，第18行表头
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import Any

import openpyxl

from app.parsers.base import BaseParser, Platform, RawRecord, ParseMeta


class WechatParser(BaseParser):
    """微信支付账单 Excel 解析器"""

    def parse(self, file_path: str) -> tuple[list[RawRecord], ParseMeta]:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        # 查找表头行
        header_row_idx = None
        for row_idx in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
            row_str = " ".join(str(v) for v in row_vals)
            if "交易时间" in row_str and "交易类型" in row_str:
                header_row_idx = row_idx
                break

        if header_row_idx is None:
            raise ValueError("无法找到微信账单表头行")

        # 提取表头
        header = [ws.cell(header_row_idx, c).value for c in range(1, ws.max_column + 1)]
        col_map = {}
        for i, h in enumerate(header):
            if h:
                col_map[str(h).strip()] = i + 1  # 1-indexed

        # 提取元数据
        meta = self._extract_metadata(ws, header_row_idx)

        # 解析数据行
        records: list[RawRecord] = []
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            record = self._parse_row(ws, col_map, row_idx)
            if record:
                records.append(record)

        wb.close()
        meta.total_count = len(records)
        return records, meta

    def _extract_metadata(self, ws, header_row_idx: int) -> ParseMeta:
        meta = ParseMeta()
        for row_idx in range(1, header_row_idx + 1):
            row_vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
            row_str = " ".join(str(v) for v in row_vals)

            # 提取昵称
            m = re.search(r"微信昵称[：:]\s*\[?(.+?)\]?", row_str)
            if m:
                meta.user_identifier = m.group(1).strip()

            # 提取时间范围
            m = re.search(r"起始时间[：:]\s*\[?(.+?)\]?", row_str)
            if m:
                meta.period_start = m.group(1).strip()
            m = re.search(r"终止时间[：:]\s*\[?(.+?)\]?", row_str)
            if m:
                meta.period_end = m.group(1).strip()

        return meta

    def _parse_row(self, ws, col_map: dict[str, int], row_idx: int) -> RawRecord | None:
        def get_val(col_name: str) -> Any:
            idx = col_map.get(col_name)
            if idx:
                return ws.cell(row_idx, idx).value
            return None

        trans_time = get_val("交易时间")
        if trans_time is None:
            return None

        direction = get_val("收/支")
        if direction not in ("支出", "收入", "不计收支"):
            direction = "支出"

        amount = get_val("金额(元)")
        if amount is None:
            return None
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            return None

        if amount <= 0:
            return None

        # 处理 datetime 对象
        if isinstance(trans_time, datetime):
            trans_dt = trans_time
        else:
            try:
                trans_dt = datetime.strptime(str(trans_time)[:19], "%Y-%m-%d %H:%M:%S")
            except (ValueError, TypeError):
                trans_dt = datetime.now()

        return RawRecord(
            platform=Platform.WECHAT,
            trans_time=trans_dt,
            amount=amount,
            direction=direction,
            merchant=str(get_val("交易对方") or ""),
            product=str(get_val("商品") or ""),
            platform_order_no=str(get_val("交易单号") or ""),
            merchant_order_no=str(get_val("商户单号") or ""),
            payment_method=str(get_val("支付方式") or ""),
            status=str(get_val("当前状态") or ""),
            note=str(get_val("备注") or ""),
            extra_fields={"transaction_type": get_val("交易类型")},
        )
